from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel
import pandas as pd
import numpy as np
from io import BytesIO
import base64
import json
import os
import re
import unicodedata
import datetime as dt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List
import google.generativeai as genai

app = FastAPI()


def clean_dni(value) -> str:
    """Normalize DNI/ID: strip whitespace, uppercase, remove leading zeros.
    Handles pandas float conversion (10008.0 -> '10008').
    """
    if pd.isna(value):
        return ""
    # If it's a float that is really an integer (e.g. 10008.0), convert to int first
    if isinstance(value, float) and value == int(value):
        value = int(value)
    s = str(value).strip().upper()
    # Remove dashes and spaces (but NOT dots yet — they were handled above)
    s = s.replace("-", "").replace(" ", "")
    # If there's a trailing ".0" still (from string-typed "10008.0"), remove it
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    # If the entire string is numeric, strip leading zeros
    if s.isdigit():
        s = s.lstrip("0") or "0"
    # If the string ends with a letter but the rest is numeric (like a Spanish DNI),
    # strip leading zeros from the numeric part.
    elif len(s) > 1 and s[:-1].isdigit():
        numeric_part = s[:-1].lstrip("0") or "0"
        s = numeric_part + s[-1]
    return s


def safe_float(val):
    if pd.isna(val):
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def find_column(df, possible_names, required=False, exact=False):
    """
    Find a column in df based on a list of possible names.
    If exact=True, it checks for case-insensitive exact match.
    If exact=False, it checks if any of the possible names are a substring of the column name.
    """
    cols_upper = {c: c.upper() for c in df.columns}
    for p in possible_names:
        p_up = p.upper()
        for orig_col, up_col in cols_upper.items():
            if exact:
                if p_up == up_col:
                    return orig_col
            else:
                if p_up in up_col:
                    return orig_col
    
    if required:
        available = list(df.columns)
        raise HTTPException(
            status_code=400, 
            detail=f"No se encontró una columna válida para {possible_names[0]}. Las columnas disponibles son: {available}"
        )
    return None

# ──────────────────────────────────────────────────────────────
# ENDPOINT 1: /api/process — Upload + clean + drop_duplicates + merge
# ──────────────────────────────────────────────────────────────
@app.post("/api/process")
async def process_payroll(
    file_xrp: UploadFile = File(...),
    file_meta4: UploadFile = File(...),
):
    try:
        # ==========================================
        # 1. READ AND CLEAN XRP
        # ==========================================
        xrp_bytes = await file_xrp.read()
        df_xrp = pd.read_excel(BytesIO(xrp_bytes), skiprows=4, engine="openpyxl")
        
        # Limpieza radical de nombres de columnas
        df_xrp.columns = df_xrp.columns.astype(str).str.strip().str.replace('\n', ' ').str.replace('\r', '')

        # Buscar columnas XRP
        XRP_COL_ID = find_column(df_xrp, ["Trabajador", "DNI trabajador", "DNI"], required=True)
        XRP_COL_NOMBRE = find_column(df_xrp, ["Nombre trabajador", "Nombre"], required=True)
        XRP_COL_DEVENGOS = find_column(df_xrp, ["Total Devengado", "Devengos"], required=True)
        XRP_COL_DEDUCCIONES = find_column(df_xrp, ["Deducciones", "Retenciones", "Retenido"], required=False)
        XRP_COL_LIQUIDO = find_column(df_xrp, ["Liquido", "Neto", "Percibir"], required=False)
        XRP_COL_CONVENIO = find_column(df_xrp, ["Convenio"], required=False)

        xrp_data = pd.DataFrame()
        xrp_data["dni_clean"] = df_xrp[XRP_COL_ID].apply(clean_dni)
        xrp_data["id_xrp"] = df_xrp[XRP_COL_ID].apply(clean_dni)
        xrp_data["nombre_xrp"] = df_xrp[XRP_COL_NOMBRE].astype(str).str.strip()
        xrp_data["devengos_xrp"] = df_xrp[XRP_COL_DEVENGOS].apply(safe_float)
        
        if XRP_COL_DEDUCCIONES:
            xrp_data["deducciones_xrp"] = df_xrp[XRP_COL_DEDUCCIONES].apply(safe_float)
        else:
            xrp_data["deducciones_xrp"] = 0.0
            
        if XRP_COL_LIQUIDO:
            xrp_data["liquido_xrp"] = df_xrp[XRP_COL_LIQUIDO].apply(safe_float)
        else:
            xrp_data["liquido_xrp"] = 0.0
            
        if XRP_COL_CONVENIO:
            xrp_data["convenio_xrp"] = df_xrp[XRP_COL_CONVENIO].apply(clean_dni)
        else:
            xrp_data["convenio_xrp"] = ""

        # Eliminar vacíos y duplicados estrictamente (XRP)
        xrp_data = xrp_data[xrp_data["dni_clean"] != ""]
        xrp_data = xrp_data.drop_duplicates(subset=['dni_clean'], keep='first')


        # ==========================================
        # 2. READ AND CLEAN META4
        # ==========================================
        meta4_bytes = await file_meta4.read()
        df_meta4 = pd.read_excel(BytesIO(meta4_bytes), skiprows=3, engine="openpyxl")
        
        # Limpieza radical de nombres de columnas
        df_meta4.columns = df_meta4.columns.astype(str).str.strip().str.replace('\n', ' ').str.replace('\r', '')

        # Buscar columnas Meta4
        META4_COL_ID = find_column(df_meta4, ["Empleado", "DNI", "Trabajador"], required=True)
        META4_COL_NOMBRE = find_column(df_meta4, ["Nombre"], required=True)
        META4_COL_EMPRESA = find_column(df_meta4, ["Centro_de_Trabajo", "Empresa", "Centro"], required=False)
        META4_COL_DEVENGOS = find_column(df_meta4, ["Total_Devengos", "Bruto", "Devengos"], required=True)
        META4_COL_DEDUCCIONES = find_column(df_meta4, ["Total.Retenido", "Deducciones", "Retenciones"], required=True)
        META4_COL_LIQUIDO = find_column(df_meta4, ["Liquido", "Neto", "Percibir"], required=True)
        META4_COL_CONVENIO = find_column(df_meta4, ["id.Convenio", "Convenio"], required=False)

        meta4_data = pd.DataFrame()
        meta4_data["dni_clean"] = df_meta4[META4_COL_ID].apply(clean_dni)
        meta4_data["id_meta4"] = df_meta4[META4_COL_ID].apply(clean_dni)

        # Nombre en Meta4 (intentando juntar apellidos si existen, o usar nombre directo)
        col_ap1 = find_column(df_meta4, ["Apellido_1", "Primer Apellido"])
        col_ap2 = find_column(df_meta4, ["Apellido_2", "Segundo Apellido"])
        
        if col_ap1:
            n_part = df_meta4[col_ap1].fillna("").astype(str).str.strip() + " "
            if col_ap2:
                n_part += df_meta4[col_ap2].fillna("").astype(str).str.strip() + ", "
            else:
                n_part += ", "
            n_part += df_meta4[META4_COL_NOMBRE].fillna("").astype(str).str.strip()
            
            meta4_data["nombre_meta4"] = n_part.str.strip().str.replace(r"^,\s*", "", regex=True)
        else:
            meta4_data["nombre_meta4"] = df_meta4[META4_COL_NOMBRE].astype(str).str.strip()

        if META4_COL_EMPRESA:
            meta4_data["empresa"] = df_meta4[META4_COL_EMPRESA].fillna("").astype(str).str.strip()
        else:
            meta4_data["empresa"] = ""
            
        meta4_data["devengos_meta4"] = df_meta4[META4_COL_DEVENGOS].apply(safe_float)
        meta4_data["deducciones_meta4"] = df_meta4[META4_COL_DEDUCCIONES].apply(safe_float)
        meta4_data["liquido_meta4"] = df_meta4[META4_COL_LIQUIDO].apply(safe_float)

        if META4_COL_CONVENIO:
            meta4_data["convenio_meta4"] = df_meta4[META4_COL_CONVENIO].apply(clean_dni)
        else:
            meta4_data["convenio_meta4"] = ""

        # Eliminar vacíos y duplicados estrictamente (Meta4)
        meta4_data = meta4_data[meta4_data["dni_clean"] != ""]
        meta4_data = meta4_data.drop_duplicates(subset=['dni_clean'], keep='first')


        # ==========================================
        # 3. MERGE (OUTER)
        # ==========================================
        df_merged = pd.merge(meta4_data, xrp_data, on="dni_clean", how="outer", indicator=True)

        result_rows = []
        for _, row in df_merged.iterrows():
            merge_status = row["_merge"] 
            
            # Preferir nombre/ID de Meta4 si existe, sino usar el de XRP
            nombre = str(row.get("nombre_meta4", "") if pd.notna(row.get("nombre_meta4")) else row.get("nombre_xrp", "")).strip()
            id_emp = str(row.get("id_meta4", "") if pd.notna(row.get("id_meta4")) else row.get("id_xrp", "")).strip()
            if not id_emp:
                id_emp = str(row["dni_clean"])
            
            empresa = str(row.get("empresa", "") if pd.notna(row.get("empresa")) else "").strip()

            # Forzar NaN a 0.0 mediante safe_float
            dev_xrp = safe_float(row.get("devengos_xrp"))
            ded_xrp = safe_float(row.get("deducciones_xrp"))
            liq_xrp = safe_float(row.get("liquido_xrp"))
            
            dev_m4 = safe_float(row.get("devengos_meta4"))
            ded_m4 = safe_float(row.get("deducciones_meta4"))
            liq_m4 = safe_float(row.get("liquido_meta4"))

            diferencia = round(liq_m4 - liq_xrp, 2)
            
            conv_xrp = str(row.get("convenio_xrp", "") if pd.notna(row.get("convenio_xrp")) else "").strip()
            conv_m4 = str(row.get("convenio_meta4", "") if pd.notna(row.get("convenio_meta4")) else "").strip()
            
            # Match strictly but ignore if both are missing
            if conv_xrp == "" and conv_m4 == "":
                conv_match = ""
            elif conv_xrp == conv_m4:
                conv_match = "COINCIDE"
            else:
                conv_match = "NO COINCIDE"

            result_rows.append({
                "nombre": nombre,
                "id_empleado": id_emp,
                "empresa": empresa,
                "devengos_xrp": round(dev_xrp, 2),
                "deducciones_xrp": round(ded_xrp, 2),
                "liquido_xrp": round(liq_xrp, 2),
                "devengos_meta4": round(dev_m4, 2),
                "deducciones_meta4": round(ded_m4, 2),
                "liquido_meta4": round(liq_m4, 2),
                "diferencia": diferencia,
                "convenio_xrp": conv_xrp,
                "convenio_meta4": conv_m4,
                "convenio_match": conv_match,
                "_merge": str(merge_status)
            })

        rows_with_diff = sum(1 for r in result_rows if abs(r["diferencia"]) > 0.01)

        return JSONResponse(content={
            "data": result_rows,
            "total_rows": len(result_rows),
            "rows_with_diff": rows_with_diff,
        })

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando archivos: {str(e)}")


# ──────────────────────────────────────────────────────────────
# ENDPOINT 2: /api/generate-excel
# ──────────────────────────────────────────────────────────────
class RowData(BaseModel):
    nombre: str = ""
    id_empleado: str = ""
    empresa: str = ""
    devengos_xrp: float = 0
    deducciones_xrp: float = 0
    liquido_xrp: float = 0
    devengos_meta4: float = 0
    deducciones_meta4: float = 0
    liquido_meta4: float = 0
    diferencia: float = 0
    convenio_xrp: str = ""
    convenio_meta4: str = ""
    convenio_match: str = ""
    _merge: str = "" # ignore when exporting


class ExcelRequest(BaseModel):
    data: List[RowData]


@app.post("/api/generate-excel")
async def generate_excel(req: ExcelRequest):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativa Nóminas"

        headers = [
            "Nombre", "ID Empleado", "Empresa",
            "Devengos XRP", "Deducciones XRP", "LÍQUIDO XRP",
            "Devengos META4", "Deducciones META4", "LÍQUIDO META4",
            "DIFERENCIA", "CONVENIO XRP", "CONVENIO META4"
        ]

        # Estilos visuales
        header_fill = PatternFill(start_color="1C4CB5", end_color="1C4CB5", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        diff_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        conv_diff_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid") # Orange/Yellow
        data_font = Font(name="Calibri", size=10)
        num_fmt = "#,##0.00"

        # Títulos
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # Datos
        for row_idx, record in enumerate(req.data, 2):
            values = [
                record.nombre, record.id_empleado, record.empresa,
                record.devengos_xrp, record.deducciones_xrp, record.liquido_xrp,
                record.devengos_meta4, record.deducciones_meta4, record.liquido_meta4,
                record.diferencia,
                record.convenio_xrp, record.convenio_meta4
            ]
            has_diff = abs(record.diferencia) > 0.01
            conv_mismatch = record.convenio_match == "NO COINCIDE"

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                # Format numbers
                if isinstance(value, (int, float)) and col_idx >= 4:
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal="right")
                
                # Format background if diff in totals
                if has_diff:
                    cell.fill = diff_fill
                
                # Format background specifically for Convenio columns if they mismatch
                if conv_mismatch and col_idx in (11, 12):
                    cell.fill = conv_diff_fill

        # Ajuste dinámico col_width
        for col_idx in range(1, len(headers) + 1):
            max_len = len(headers[col_idx - 1])
            for row_idx in range(2, len(req.data) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 35)

        ws.freeze_panes = "A2"

        # Convertir a Buffer -> Base64
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode("utf-8")

        return JSONResponse(content={"excel_base64": b64})

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")


def _normalize_header(value: str) -> str:
    value = value.strip().lower()
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", value)


def _extract_target_columns(payload) -> List[str]:
    if isinstance(payload, dict):
        if "columnas" in payload:
            cols = payload["columnas"]
        elif "R0" in payload and isinstance(payload["R0"], dict) and "columnas" in payload["R0"]:
            cols = payload["R0"]["columnas"]
        else:
            cols = []
    elif isinstance(payload, list):
        cols = payload
    else:
        cols = []

    out: List[str] = []
    for c in cols:
        if isinstance(c, dict) and "name" in c:
            out.append(str(c["name"]))
        elif isinstance(c, str):
            out.append(c)
    return out


def _load_sgel_template() -> tuple[List[str], dict, dict]:
    base_dir = os.path.dirname(__file__)
    primary_path = os.path.join(base_dir, "templates", "r_formats.json")
    fallback_path = os.path.join(base_dir, "templates", "sgel_r_format.json")

    if os.path.exists(primary_path):
        with open(primary_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        r0 = payload.get("R0", {}) if isinstance(payload, dict) else {}
        columns = r0.get("columnas", []) if isinstance(r0, dict) else []
        target_cols: List[str] = []
        defaults: dict = {}
        meta_map: dict = {}
        for col in columns:
            if not isinstance(col, dict):
                continue
            name = str(col.get("name", "")).strip()
            if not name:
                continue
            target_cols.append(name)
            meta_map[name] = col
            if "default" in col:
                defaults[name] = col.get("default")
        if target_cols:
            return target_cols, defaults, meta_map

    if os.path.exists(fallback_path):
        with open(fallback_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        cols = _extract_target_columns(payload)
        if cols:
            return cols, {}, {}

    raise HTTPException(status_code=500, detail="No se encontrÃ³ un archivo de referencia SGEL vÃ¡lido.")


def _heuristic_mapping(source_headers: List[str], target_headers: List[str]) -> dict:
    source_map = { _normalize_header(h): h for h in source_headers }
    mapping = {}
    for target in target_headers:
        t_norm = _normalize_header(target)
        if t_norm in source_map:
            mapping[target] = source_map[t_norm]
            continue
        # Partial contains matching as fallback
        found = ""
        for s_norm, s_raw in source_map.items():
            if t_norm and (t_norm in s_norm or s_norm in t_norm):
                found = s_raw
                break
        mapping[target] = found
    return mapping


def _detect_header_row(df_raw: pd.DataFrame, max_rows: int = 20) -> int:
    best_row = 0
    best_score = -1
    rows_to_check = min(max_rows, len(df_raw))
    for i in range(rows_to_check):
        row = df_raw.iloc[i].tolist()
        non_empty = [c for c in row if isinstance(c, str) and c.strip()]
        score = len(non_empty)
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _is_date_like(value) -> bool:
    if pd.isna(value):
        return False
    return isinstance(value, (pd.Timestamp, dt.date, dt.datetime))


def _normalize_name_for_checks(value: str) -> str:
    return _normalize_header(value)


def _map_headers_with_gemini(source_headers: List[str], target_headers: List[str]) -> dict:
    api_key = os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="Missing GOOGLE_API_KEY or GEMINI_API_KEY for Gemini.")

    genai.configure(api_key=api_key)
    model_candidates = []
    env_model = os.getenv("GEMINI_MODEL")
    if env_model:
        model_candidates.append(env_model)
    model_candidates.extend([
        "gemini-1.5-flash",
        "gemini-1.5-pro",
        "gemini-1.0-pro",
    ])

    prompt = (
        "You map Excel headers to target headers.\n"
        "Return ONLY valid JSON: {\"TargetColumn\": \"SourceColumn\"}.\n"
        "If no match, return empty string.\n\n"
        f"Source headers: {source_headers}\n"
        f"Target headers: {target_headers}\n"
    )

    text = ""
    last_error = None
    for model_name in model_candidates:
        try:
            model = genai.GenerativeModel(model_name)
            response = model.generate_content(prompt)
            text = response.text or ""
            if text:
                break
        except Exception as e:
            last_error = e
            continue

    if not text:
        if last_error:
            print(f"[gemini] model error: {last_error}")
        return {}

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        return {}

    try:
        data = json.loads(text[start : end + 1])
    except json.JSONDecodeError:
        return {}

    if not isinstance(data, dict):
        return {}
    return {str(k): (str(v) if v is not None else "") for k, v in data.items()}


@app.post("/api/generate-sgel-r")
async def generate_sgel_r(file: UploadFile = File(...)):
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="Archivo no vÃ¡lido.")

        target_columns, defaults, meta_map = _load_sgel_template()

        raw = await file.read()
        df_raw = pd.read_excel(BytesIO(raw), engine="openpyxl", header=None)
        header_row = _detect_header_row(df_raw)
        headers = df_raw.iloc[header_row].tolist()
        headers = [str(h).strip() if pd.notna(h) else "" for h in headers]
        df = df_raw.iloc[header_row + 1 :].copy()
        df.columns = headers
        df = df.loc[:, [c for c in df.columns if c]]
        df = df.dropna(how="all")
        source_headers = [str(c) for c in df.columns]
        print(f"[sgel] detected header row: {header_row}")
        print(f"[sgel] source headers: {source_headers}")
        print(f"[sgel] source rows: {len(df)}")
        if len(df) > 0:
            print(f"[sgel] first row sample: {df.iloc[0].to_dict()}")

        mapping = _map_headers_with_gemini(source_headers, target_columns)
        if mapping:
            print(f"[sgel] gemini mapping: {mapping}")
        if not mapping:
            mapping = _heuristic_mapping(source_headers, target_columns)
            print(f"[sgel] heuristic mapping: {mapping}")

        normalized_sources = { _normalize_header(c): c for c in df.columns }

        out = pd.DataFrame(index=df.index)
        for target in target_columns:
            if target in defaults:
                out[target] = defaults[target]
                continue
            source = mapping.get(target, "")
            if source in df.columns:
                out[target] = df[source]
            else:
                alt = normalized_sources.get(_normalize_header(source), "")
                if alt in df.columns:
                    out[target] = df[alt]
                elif target in defaults:
                    out[target] = defaults[target]
                else:
                    out[target] = pd.NA

        issues = []
        for col in out.columns:
            meta = meta_map.get(col, {})
            expected_type = str(meta.get("type", "string")).lower()
            series = out[col]

            if expected_type == "date":
                parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
                invalid = series.notna() & parsed.isna()
                for idx, val in series[invalid].items():
                    issues.append({
                        "fila": int(idx) + 1,
                        "columna": col,
                        "valor": val,
                        "motivo": "Fecha no valida"
                    })
                fmt = meta.get("format", "%d/%m/%Y")
                out[col] = parsed.dt.strftime(fmt)
            elif expected_type == "number":
                num = pd.to_numeric(series, errors="coerce")
                invalid = series.notna() & num.isna()
                for idx, val in series[invalid].items():
                    issues.append({
                        "fila": int(idx) + 1,
                        "columna": col,
                        "valor": val,
                        "motivo": "Numero no valido"
                    })
                out[col] = num
            else:
                name_norm = _normalize_name_for_checks(col)
                if "fecha" not in name_norm:
                    invalid = series.apply(_is_date_like)
                    for idx, val in series[invalid].items():
                        issues.append({
                            "fila": int(idx) + 1,
                            "columna": col,
                            "valor": val,
                            "motivo": "Valor de fecha en columna no fecha"
                        })
                    out.loc[invalid, col] = pd.NA
        if len(out) > 0:
            print(f"[sgel] output first row sample: {out.iloc[0].to_dict()}")

        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            out.to_excel(writer, index=False, sheet_name="Plantilla R")
            if issues:
                pd.DataFrame(issues).to_excel(writer, index=False, sheet_name="Validaciones")
        buf.seek(0)

        headers = {
            "Content-Disposition": "attachment; filename=Plantilla_R_Resultado.xlsx"
        }
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando SGEL R: {str(e)}")
        
