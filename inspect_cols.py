import openpyxl

print("=== XRP FILE — ALL HEADERS (Row 5) ===")
wb = openpyxl.load_workbook('fh_xrp.xlsx', read_only=True, data_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=5, max_row=5, values_only=True):
    cols = [c for c in row if c is not None]
    for i, c in enumerate(cols):
        print(f"  [{i}] {c}")
wb.close()

print("\n=== META4 FILE — ALL HEADERS (Row 4) ===")
wb2 = openpyxl.load_workbook('fh_meta4.xlsx', read_only=True, data_only=True)
ws2 = wb2.active
for row in ws2.iter_rows(min_row=4, max_row=4, values_only=True):
    cols = [c for c in row if c is not None]
    for i, c in enumerate(cols):
        print(f"  [{i}] {c}")
wb2.close()
